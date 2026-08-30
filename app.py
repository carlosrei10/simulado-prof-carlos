import json
import os
import re
import time
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from google import genai
from google.genai import types
from google.genai.errors import APIError
import streamlit as st
import streamlit.components.v1 as components

# -----------------------------------------------------------------------------
# CONFIGURAÇÃO DA PÁGINA
# -----------------------------------------------------------------------------
st.set_page_config(
    page_title="👑 Rei dos Simulados",
    page_icon="👑",
    layout="centered",
    initial_sidebar_state="collapsed",
)

MODEL_NAME = "gemini-3.6-flash"

# -----------------------------------------------------------------------------
# INICIALIZAÇÃO DA API GEMINI
# -----------------------------------------------------------------------------
try:
    api_key = st.secrets.get("GEMINI_API_KEY", os.environ.get("GEMINI_API_KEY", ""))
except Exception:
    api_key = os.environ.get("GEMINI_API_KEY", "")

if not api_key:
    api_key = "AQ.Ab8RN6LjVTB0KTK4HsQb9F-oCzn2qeiur5Uga5YxZ6ietrw56w"

client = genai.Client(api_key=api_key)

# -----------------------------------------------------------------------------
# ESTILIZAÇÃO CSS GLOBAL (CORREÇÃO DE OPACIDADE E TEXTO AO CLICAR/FOCO)
# -----------------------------------------------------------------------------
bg_body = "#F0F4F8"
text_color = "#0F172A"
card_bg = "#FFFFFF"
card_border = "#CBD5E1"
header_gradient = "linear-gradient(135deg, #0B192C 0%, #1E3A8A 50%, #2563EB 100%)"
hero_bg = "linear-gradient(135deg, #0F172A 0%, #1E3A8A 50%, #2563EB 100%)"
btn_bg = "#2563EB"

st.markdown(
    f"""
    <style>
        @import url('https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@400;600;700;800;900&display=swap');

        * {{
            font-family: 'Plus Jakarta Sans', sans-serif;
        }}

        .stApp {{
            background-color: {bg_body} !important;
            color: {text_color} !important;
        }}

        .block-container {{
            max-width: 720px !important;
            padding-top: 1.5rem !important;
            padding-bottom: 1.5rem !important;
            margin: 0 auto !important;
        }}

        /* HERO BANNER LANDING PAGE */
        .hero-banner {{
            background: {hero_bg};
            border-radius: 16px;
            padding: 20px 24px;
            text-align: center;
            color: white;
            box-shadow: 0 8px 20px rgba(0, 0, 0, 0.2);
            margin-bottom: 20px;
            position: relative;
            overflow: hidden;
            border: 1px solid rgba(255, 255, 255, 0.15);
        }}

        .hero-coroa {{
            font-size: 2.2rem;
            margin-bottom: 4px;
            display: block;
            filter: drop-shadow(0 2px 4px rgba(0,0,0,0.3));
        }}

        .hero-titulo {{
            font-size: 1.9rem !important;
            font-weight: 900 !important;
            letter-spacing: -0.5px;
            margin: 0 !important;
            text-transform: uppercase;
            background: linear-gradient(180deg, #FFFFFF 0%, #E2E8F0 100%);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
        }}

        .hero-subtitulo {{
            font-size: 0.95rem !important;
            font-weight: 700 !important;
            color: #FACC15 !important;
            margin-top: 2px !important;
            margin-bottom: 10px !important;
            letter-spacing: 0.5px;
            text-transform: uppercase;
        }}

        .hero-bordao {{
            font-size: 0.85rem !important;
            font-weight: 500;
            color: rgba(255, 255, 255, 0.95);
            max-width: 480px;
            margin: 0 auto !important;
            line-height: 1.3;
            background: rgba(255, 255, 255, 0.1);
            padding: 6px 16px;
            border-radius: 20px;
            backdrop-filter: blur(5px);
            display: inline-block;
        }}

        /* CAMPOS DE SELEÇÃO E TEXTO */
        div[data-baseweb="select"], .stSelectbox > div > div {{
            background-color: #FFFFFF !important;
            border: 2px solid #CBD5E1 !important;
            border-radius: 10px !important;
            color: #0F172A !important;
            box-shadow: 0 2px 6px rgba(0,0,0,0.05) !important;
        }}

        div[data-baseweb="select"] * {{
            font-size: 0.95rem !important;
            font-weight: 700 !important;
            color: #0F172A !important;
        }}

        input[data-baseweb="input"], div[data-baseweb="input"], .stTextInput > div > div {{
            background-color: #FFFFFF !important;
            border: 2px solid #CBD5E1 !important;
            border-radius: 10px !important;
            color: #0F172A !important;
            font-size: 0.95rem !important;
            font-weight: 600 !important;
            box-shadow: 0 2px 6px rgba(0,0,0,0.05) !important;
        }}

        .rotulo-seletor {{
            font-size: 0.95rem !important;
            font-weight: 800 !important;
            color: {text_color} !important;
            margin-bottom: 6px !important;
            display: flex;
            align-items: center;
            gap: 6px;
            white-space: nowrap;
        }}

        /* BOTÕES GERAIS E CORREÇÃO DE OPACIDADE AO CLICAR/FOCO */
        div.stButton > button {{
            font-size: 0.95rem;
            padding: 12px 14px;
            background-color: {btn_bg};
            color: #FFFFFF !important;
            border-radius: 10px;
            border: none;
            font-weight: 700;
            text-align: center;
            justify-content: center;
            line-height: 1.4;
            white-space: normal;
            word-wrap: break-word;
            margin-bottom: 6px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.08);
            opacity: 1 !important;
        }}

        div.stButton > button:hover, 
        div.stButton > button:focus, 
        div.stButton > button:active {{
            opacity: 1 !important;
            color: #FFFFFF !important;
            background-color: #1D4ED8 !important;
            border-color: transparent !important;
            outline: none !important;
            box-shadow: 0 2px 8px rgba(0,0,0,0.15) !important;
        }}

        div.stButton > button p, 
        div.stButton > button span {{
            color: #FFFFFF !important;
            opacity: 1 !important;
            font-weight: 700 !important;
        }}

        /* CARDS LANDING PAGE */
        div.stButton > button[key="btn_basica"], div.stButton > button[key="btn_superior"] {{
            background-color: {btn_bg} !important;
            color: #FFFFFF !important;
            border-radius: 0px 0px 14px 14px !important;
            margin-top: -2px !important;
            width: 100% !important;
            text-align: center !important;
            justify-content: center !important;
            font-weight: 800 !important;
            font-size: 0.95rem !important;
            padding: 12px 14px !important;
            opacity: 1 !important;
        }}

        .card-basica-topo, .card-superior-topo {{
            background-color: {card_bg};
            border: 2px solid {card_border};
            border-bottom: none;
            border-radius: 14px 14px 0px 0px;
            padding: 16px 14px 10px 14px;
            text-align: center;
            box-shadow: 0 4px 12px rgba(0,0,0,0.05);
        }}

        .card-modulo {{
            background-color: {card_bg};
            border: 2px solid {card_border};
            border-radius: 14px;
            padding: 16px 12px;
            text-align: center;
            box-shadow: 0px 4px 10px rgba(0,0,0,0.05);
            margin-bottom: 8px;
            min-height: 160px;
        }}

        .enunciado-grande {{
            font-size: 1.05rem !important;
            font-weight: 700;
            line-height: 1.5;
            color: {text_color};
            margin-bottom: 16px;
            padding: 16px;
            background-color: {card_bg};
            border: 2px solid {card_border};
            border-left: 6px solid {btn_bg};
            border-radius: 10px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.04);
        }}

        .card-dificuldade {{
            border-radius: 10px;
            padding: 10px 12px;
            text-align: center;
            color: white;
            font-weight: bold;
            margin-top: 4px;
            font-size: 0.95rem;
            display: flex;
            align-items: center;
            justify-content: center;
            gap: 8px;
        }}

        .card-resultado {{
            background-color: {card_bg};
            border: 2px solid {card_border};
            border-radius: 14px;
            text-align: center;
        }}

        .metric-box {{
            background-color: rgba(0,0,0,0.03);
            padding: 12px;
            border-radius: 8px;
            border: 1px solid {card_border};
        }}
    </style>
    """,
    unsafe_allow_html=True,
)

# -----------------------------------------------------------------------------
# GERENCIAMENTO DE ESTADO (SESSION STATE)
# -----------------------------------------------------------------------------
if "etapa_ensino" not in st.session_state:
    st.session_state.etapa_ensino = None
if "funcao_selecionada" not in st.session_state:
    st.session_state.funcao_selecionada = None
if "questoes_online" not in st.session_state:
    st.session_state.questoes_online = []
if "respostas_usuario" not in st.session_state:
    st.session_state.respostas_usuario = {}
if "indice_questao" not in st.session_state:
    st.session_state.indice_questao = 0
if "qtd_total_questoes" not in st.session_state:
    st.session_state.qtd_total_questoes = 10
if "tempo_inicio" not in st.session_state:
    st.session_state.tempo_inicio = None
if "simulado_concluido" not in st.session_state:
    st.session_state.simulado_concluido = False


def formatar_tempo(segundos):
    minutos = int(segundos // 60)
    segundos_restantes = int(segundos % 60)
    return f"{minutos:02d}:{segundos_restantes:02d}"


def rolar_para_o_topo():
    components.html(
        """
        <script>
            window.parent.document.querySelector('.main').scrollTo({top: 0, behavior: 'smooth'});
        </script>
        """,
        height=0,
    )


def limpar_e_formatar_texto(texto):
    if not texto:
        return ""
    texto = re.sub(r"`([^`]+)`", r"\1", texto)
    texto = texto.replace("$", "")
    texto = re.sub(r"\\frac\{([^}]+)\}\{([^}]+)\}", r"\1/\2", texto)
    texto = re.sub(r"\\text\{([^}]+)\}", r"\1", texto)
    texto = texto.replace(r"\quad", " ")
    texto = texto.replace(r"\times", "×")
    texto = texto.replace(r"\div", "÷")
    texto = texto.replace("\\", "")
    return texto


def resetar_para_inicio():
    st.session_state.etapa_ensino = None
    st.session_state.funcao_selecionada = None
    st.session_state.questoes_online = []
    st.session_state.respostas_usuario = {}
    st.session_state.indice_questao = 0
    st.session_state.tempo_inicio = None
    st.session_state.simulado_concluido = False


def voltar_etapa():
    if st.session_state.simulado_concluido:
        st.session_state.simulado_concluido = False
    elif st.session_state.questoes_online:
        st.session_state.questoes_online = []
        st.session_state.respostas_usuario = {}
        st.session_state.indice_questao = 0
    elif st.session_state.funcao_selecionada:
        st.session_state.funcao_selecionada = None
    elif st.session_state.etapa_ensino:
        st.session_state.etapa_ensino = None


def renderizar_botoes_navegacao(posicao_key):
    col_v, col_i = st.columns(2)
    with col_v:
        if st.button("⬅️ Voltar", key=f"btn_voltar_{posicao_key}", use_container_width=True):
            voltar_etapa()
            st.rerun()
    with col_i:
        if st.button("🏠 Início", key=f"btn_inicio_{posicao_key}", use_container_width=True):
            resetar_para_inicio()
            st.rerun()


def gerar_lote_questoes(materia, topico, etapa, quantidade, nivel):
    descricoes_nivel = {
        1: "Nível 1 (Fácil / Conceitual): Questões diretas de fixação de conceitos básicos.",
        2: "Nível 2 (Tranquilo / Prático): Questões simples com cenários práticos do dia a dia.",
        3: "Nível 3 (Moderado / Acadêmico): Questões no padrão tradicional de exames e provas escolares/acadêmicas.",
        4: "Nível 4 (Desafiador / Análise): Questões complexas com casos práticos e pegadinhas.",
        5: "Nível 5 (Nível Concurso Público / Professor): Questões no estilo exato de bancas examinadoras de Concursos Públicos (Cebraspe, FGV, Vunesp, FCC). Altamente exigente.",
    }

    system_instruction = (
        "Você é um renomado examinador de provas e concursos públicos do Magistério. "
        "Sua regra absoluta é respeitar estritamente o tema solicitado e o nível de dificuldade exigido."
    )

    prompt_json = (
        f"Gere um array com exatamente {quantidade} questões inéditas para a disciplina {materia} ({etapa}).\n\n"
        f"🎯 TEMA EXCLUSIVO: \"{topico}\"\n"
        f"📊 DIFICULDADE EXIGIDA: {descricoes_nivel[nivel]}\n\n"
        "REGRAS DE FORMATAÇÃO:\n"
        "Retorne EXCLUSIVAMENTE um JSON com uma lista sob a chave \"questoes\", no formato:\n"
        "{\n"
        '  "questoes": [\n'
        "    {\n"
        '      "enunciado": "Texto do enunciado",\n'
        '      "alternativas": {"a": "Opção A", "b": "Opção B", "c": "Opção C", "d": "Opção D", "e": "Opção E"},\n'
        '      "correta": "a",\n'
        '      "explicacao": "Explicação objetiva em até 2 frases."\n'
        "    }\n"
        "  ]\n"
        "}"
    )

    max_tentativas = 3
    for tentativa in range(max_tentativas):
        try:
            res = client.models.generate_content(
                model=MODEL_NAME,
                contents=prompt_json,
                config=types.GenerateContentConfig(
                    system_instruction=system_instruction,
                    response_mime_type="application/json",
                    temperature=0.3,
                ),
            )
            dados = json.loads(res.text)
            return dados.get("questoes", [])
        except APIError as e:
            if "429" in str(e) or "RESOURCE_EXHAUSTED" in str(e):
                if tentativa < max_tentativas - 1:
                    time.sleep(11)
                    continue
                else:
                    st.error("⚠️ Limite de requisições atingido. Aguarde cerca de 10 segundos antes de clicar novamente.")
                    return []
            else:
                st.error(f"Erro ao comunicar com a API Gemini: {e}")
                return []
        except Exception as e:
            st.error(f"Erro no processamento da resposta: {e}")
            return []


# -----------------------------------------------------------------------------
# FUNÇÃO DE EXPORTAÇÃO PARA EXCEL (OPENPYXL)
# -----------------------------------------------------------------------------
def criar_arquivo_excel(questoes):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Simulado - Questões"
    ws.views.sheetView[0].showGridLines = True

    headers = [
        "Pergunta / Enunciado", 
        "Alternativa A", 
        "Alternativa B", 
        "Alternativa C", 
        "Alternativa D", 
        "Alternativa E", 
        "Comentário da Resposta"
    ]
    ws.append(headers)

    for q in questoes:
        alt = q.get("alternativas", {})
        ws.append([
            q.get("enunciado", ""),
            alt.get("a", ""),
            alt.get("b", ""),
            alt.get("c", ""),
            alt.get("d", ""),
            alt.get("e", ""),
            q.get("explicacao", "")
        ])

    header_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'), right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'), bottom=Side(style='thin', color='D3D3D3')
    )
    zebra_fill = PatternFill(start_color="F9FAFC", end_color="F9FAFC", fill_type="solid")

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28

    for row_num in range(2, len(questoes) + 2):
        ws.row_dimensions[row_num].height = 65
        is_even = (row_num % 2 == 0)
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.font = Font(name="Calibri", size=10)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            if is_even:
                cell.fill = zebra_fill

    column_widths = {'A': 45, 'B': 28, 'C': 28, 'D': 28, 'E': 28, 'F': 28, 'G': 45}
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    ws.freeze_panes = 'A2'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# -----------------------------------------------------------------------------
# ESTRUTURA DE CONTEÚDOS (EM ORDEM ALFABÉTICA)
# -----------------------------------------------------------------------------

CONTEUDOS_EDUCACAO_BASICA = {
    "Artes": {
        "Artes Visuais e História da Arte": [
            "Elementos da Linguagem Visual (Ponto, Linha, Forma, Cor)",
            "Arte Pré-Histórica e Antiga",
            "Movimentos Artísticos Modernos e Contemporâneos",
            "Arte Brasileira e Patrimônio Cultural",
        ]
    },
    "Biologia": {
        "Citologia e Genética": [
            "Estrutura e Função Celular",
            "Divisão Celular (Mitose e Meiose)",
            "Leis de Mendel e Genética Humana",
            "Biotecnologia e Engenharia Genética",
        ],
        "Ecologia e Evolução": [
            "Cadeias e Teias Alimentares",
            "Relações Ecológicas e Biomas Brasileiros",
            "Teorias Evolutivas (Lamarckismo, Darwinismo, Neodarwinismo)",
        ],
    },
    "Ciências": {
        "Corpo Humano e Meio Ambiente": [
            "Sistemas do Corpo Humano e Saúde",
            "Matéria, Energia e Misturas",
            "Terra, Universo e Sistema Solar",
            "Preservação Ambiental e Sustentabilidade",
        ]
    },
    "Educação Física": {
        "Cultura Corporal do Movimento": [
            "Jogos, Esportes e Brincadeiras",
            "Ginástica, Dança e Lutas",
            "Saúde, Anatomia e Qualidade de Vida",
        ]
    },
    "Filosofia": {
        "História do Pensamento Filosófico": [
            "Filosofia Antiga (Sócrates, Platão, Aristóteles)",
            "Filosofia Política e Ética",
            "Teoria do Conhecimento (Racionalismo e Empirismo)",
            "Filosofia Contemporânea",
        ]
    },
    "Física": {
        "Mecânica e Termodinâmica": [
            "Cinemática (Movimento Uniforme e Variado)",
            "Leis de Newton e Dinâmica",
            "Trabalho, Energia e Potência",
            "Termologia e Calorimetria",
        ],
        "Eletromagnetismo e Óptica": [
            "Eletrostática e Circuitos Elétricos",
            "Ondulatória e Fenômenos Ondulatórios",
            "Óptica Geométrica (Lentes e Espelhos)",
        ],
    },
    "Geografia": {
        "Geografia Física e Humana": [
            "Cartografia, Fuso Horário e Orientação",
            "Relevo, Clima e Hidrografia",
            "Geografia Urbana e Agrária",
            "Geopolítica e Globalização",
            "Geografia do Brasil",
        ]
    },
    "História": {
        "História Geral e do Brasil": [
            "Antiguidade Clássica (Grécia e Roma)",
            "Idade Média e Feudalismo",
            "Brasil Colônia, Império e República",
            "Primeira e Segunda Guerra Mundial",
            "Guerra Fria e Mundo Contemporâneo",
        ]
    },
    "Língua Espanhola": {
        "Linguagem e Compreensão": [
            "Compreensão e Interpretação Textual em Espanhol",
            "Gramática, Pronomes e Conectores",
            "Falsos Amigos (Heterosemânticos)",
        ]
    },
    "Língua Inglesa": {
        "Linguagem e Compreensão": [
            "Reading Comprehension & Vocabulary",
            "Grammar & Verb Tenses",
            "Linking Words & Textual Cohesion",
        ]
    },
    "Língua Portuguesa": {
        "Gramática e Interpretação": [
            "Compreensão e Interpretação de Textos",
            "Ortografia, Acentuação e Pontuação",
            "Classes de Palavras e Sintaxe",
            "Concordância, Regência e Crase",
            "Literatura Brasileira e Portuguesa",
        ]
    },
    "Matemática": {
        "Álgebra, Geometria e Estatística": [
            "Operações Fundamentais e Frações",
            "Equações de 1º e 2º Graus e Sistemas",
            "Porcentagem, Regra de Três e Juros",
            "Geometria Plana e Espacial",
            "Funções, Estatística e Probabilidade",
        ]
    },
    "Química": {
        "Química Geral e Orgânica": [
            "Estrutura Atômica e Tabela Periódica",
            "Ligações Químicas e Funções Inorgânicas",
            "Estequiometria e Soluções",
            "Química Orgânica (Cadeias e Funções)",
        ]
    },
    "Sociologia": {
        "Sociedade e Indivíduo": [
            "Clássicos da Sociologia (Durkheim, Marx, Weber)",
            "Cultura, Identidade e Socialização",
            "Trabalho, Desigualdade e Cidadania",
        ]
    },
}

CONTEUDOS_ENSINO_SUPERIOR = {
    "Conhecimentos Pedagógicos": {
        "Didática e Prática de Ensino": [
            "Tendências Pedagógicas (Liberal, Progressista, Libertadora)",
            "Planejamento Escolar (Anual, de Ensino e Plano de Aula)",
            "Projeto Político-Pedagógico (PPP)",
            "Avaliação Escolar (Diagnóstica, Formativa e Somativa)",
            "Processo de Ensino-Aprendizagem e Relação Professor-Aluno",
            "Gestão Democrática e Conselho Escolar",
            "Metodologias Ativas e Sala de Aula Invertida",
            "Interdisciplinaridade e Transdisciplinaridade",
        ],
        "Psicologia e Teoria da Educação": [
            "Teoria do Desenvolvimento de Jean Piaget",
            "Sociointeracionismo de Vygotsky (ZDP)",
            "Psicogenética de Henri Wallon",
            "Teoria das Inteligências Múltiplas (Gardner)",
            "Psicologia da Aprendizagem e Cognição",
            "Educação Inclusiva e Necessidades Educacionais Especiais (NEE)",
            "Bullying e Convivência Escolar",
        ],
        "Currículo e Sociedade": [
            "Teorias do Currículo (Tradicional, Crítica e Pós-Crítica)",
            "Diversidade Cultural e Relações Etnico-Raciais",
            "Tecnologias Digitais da Informação e Comunicação (TDIC)",
            "Educação Integral e Tempo Integral",
        ],
    },
    "Legislação Educacional": {
        "Leis Federais Fundamentais": [
            "LDB - Lei 9.394/1996 (Princípios, Fins e Organização)",
            "LDB - Educação Infantil, Ensino Fundamental e Médio",
            "ECA - Estatuto da Criança e do Adolescente (Direito à Educação)",
            "Constituição Federal de 1988 (Artigos 205 ao 214)",
            "LBI - Lei Brasileira de Inclusão (Lei nº 13.146/2015)",
            "FUNDEB - Lei nº 14.113/2020",
        ],
        "Diretrizes, BNCC e Planos": [
            "BNCC - Estrutura, Competências Gerais e Áreas do Conhecimento",
            "BNCC na Educação Infantil e Ensino Fundamental",
            "PNE - Plano Nacional de Educação (Metas e Diretrizes)",
            "DCNs - Diretrizes Curriculares Nacionais para a Educação Básica",
            "Diretrizes para a Educação das Relações Étnico-Raciais",
        ],
    },
    "Língua Portuguesa para Concurso": {
        "Gramática e Ortografia": [
            "Nova Ortografia e Acentuação Gráfica",
            "Classes de Palavras (Substantivo, Adjetivo, Verbo, etc.)",
            "Sintaxe da Oração e do Período",
            "Concordância Verbal e Nominal",
            "Regência Verbal, Nominal e Emprego da Crase",
            "Pontuação e Seus Usos Expressivos",
        ],
        "Interpretação e Coesão": [
            "Compreensão e Interpretação de Textos",
            "Mecanismos de Coesão e Coerência Textual",
            "Tipologia e Gêneros Textuais",
            "Figuras de Linguagem e Funções da Linguagem",
        ],
    },
    "Raciocínio Lógico e Matemática": {
        "Lógica Proposicional": [
            "Proposições Simples e Compostas",
            "Conectivos Lógicos (E, OU, SE...ENTÃO, SE E SOMENTE SE)",
            "Tabelas-Verdade, Tautologia, Contradição e Contingência",
            "Negação de Proposições e Equivalências Lógicas",
            "Argumentação Lógica e Validade de Argumentos",
        ],
        "Lógica Quantitativa e Matemática": [
            "Diagramas Lógicos, Conjuntos e Operações",
            "Sequências Lógicas, Numéricas e Figurais",
            "Princípio da Casa dos Pombos (Princípio da Gaveta)",
            "Análise Combinatória (Arranjo, Permutação e Combinação)",
            "Probabilidade Simples e Condicional",
            "Porcentagem, Juros Simples e Regra de Três",
        ],
    },
}

# -----------------------------------------------------------------------------
# LANDING PAGE CENTRALIZADA (TELA PRINCIPAL)
# -----------------------------------------------------------------------------
if st.session_state.etapa_ensino is None:
    st.markdown(
        """
        <div class="hero-banner">
            <span class="hero-coroa">👑</span>
            <h1 class="hero-titulo">REI DOS SIMULADOS</h1>
            <div class="hero-subtitulo">⚡ TREINE COM PRECISÃO ⚡</div>
            <div class="hero-bordao">"Domine a banca, treine com precisão e conquiste a sua vaga!"</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    st.markdown(
        f"<h3 style='text-align: center; color: {text_color}; font-weight: 800; margin-bottom: 16px; font-size: 1.15rem; text-transform: uppercase;'>SELECIONE A SUA ETAPA DE ESTUDOS:</h3>",
        unsafe_allow_html=True,
    )

    col1, col2 = st.columns(2)

    with col1:
        st.markdown(
            f"""
            <div class="card-basica-topo">
                <div style="font-size: 1.2rem; font-weight: 800; color: {btn_bg}; margin-bottom: 4px;">🏫 EDUCAÇÃO BÁSICA</div>
                <div style="font-size: 0.85rem; color: {text_color}; opacity: 0.85; margin-bottom: 8px;">Ensino Fundamental e Médio • ENEM • Provas Escolares</div>
                <div style="font-size: 1.4rem;">📐 📖 🎨 🧪 🌍</div>
            </div>
            """,
            unsafe_allow_html=True,
        )
        if st.button("👈 ACESSAR EDUCAÇÃO BÁSICA 👉", key="btn_basica", use_container_width=True):
            st.session_state.etapa_ensino = "Educação Básica"
            st.rerun()

    with col2:
        st.markdown(
            f"""
            <div class="card-superior-topo">
                <div style="font-size: 1.2rem; font-weight: 800; color: {btn_bg}; margin-bottom: 4px;">🎓 ENSINO SUPERIOR</div>
                <div style="font-size: 0.85rem; color: {text_color}; opacity: 0.85; margin-bottom: 8px;">Concursos de Magistério & Nível Superior</div>
                <div style="font-size: 1.4rem;">📜 🧠 📖 🔢</div>
            </div>
            """,
            unsafe_allow_html=True,
        )
        if st.button("👈 ACESSAR ENSINO SUPERIOR 👉", key="btn_superior", use_container_width=True):
            st.session_state.etapa_ensino = "Ensino Superior"
            st.rerun()

# -----------------------------------------------------------------------------
# PAINEL DE SELEÇÃO DE MATÉRIA E GERAÇÃO DE QUESTÕES
# -----------------------------------------------------------------------------
elif st.session_state.etapa_ensino and not st.session_state.questoes_online and not st.session_state.simulado_concluido:
    st.markdown(
        f"""
        <div class="hero-banner" style="padding: 14px 20px; margin-bottom: 15px;">
            <h2 style="font-size: 1.4rem; font-weight: 900; margin: 0; text-transform: uppercase;">{st.session_state.etapa_ensino}</h2>
            <p style="font-size: 0.85rem; margin: 4px 0 0 0; color: #FACC15;">Configure o seu simulado personalizado</p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    base_conteudos = CONTEUDOS_EDUCACAO_BASICA if st.session_state.etapa_ensino == "Educação Básica" else CONTEUDOS_ENSINO_SUPERIOR

    materias_disponiveis = sorted(list(base_conteudos.keys()))
    materia_escolhida = st.selectbox("📚 Selecione a Disciplina:", materias_disponiveis)

    subcategorias = base_conteudos[materia_escolhida]
    subcat_escolhida = st.selectbox("📂 Selecione o Eixo / Subcategoria:", sorted(list(subcategorias.keys())))

    topicos_disponiveis = subcategorias[subcat_escolhida]
    topico_escolhido = st.selectbox("🎯 Selecione o Tópico Específico:", topicos_disponiveis)

    col_qtd, col_niv = st.columns(2)
    with col_qtd:
        quantidade_questoes = st.selectbox("🔢 Quantidade de Questões:", [5, 10, 15, 20], index=1)
    with col_niv:
        nivel_dificuldade = st.selectbox("⭐ Nível de Dificuldade:", [1, 2, 3, 4, 5], index=2, format_func=lambda x: f"Nível {x}")

    st.write("")
    if st.button("🚀 GERAR SIMULADO AGORA", use_container_width=True):
        with st.spinner("✨ Gerando questões inéditas com Inteligência Artificial..."):
            questoes_geradas = gerar_lote_questoes(
                materia=materia_escolhida,
                topico=topico_escolhido,
                etapa=st.session_state.etapa_ensino,
                quantidade=quantidade_questoes,
                nivel=nivel_dificuldade
            )
            if questoes_geradas:
                st.session_state.questoes_online = questoes_geradas
                st.session_state.respostas_usuario = {}
                st.session_state.indice_questao = 0
                st.session_state.tempo_inicio = time.time()
                st.session_state.simulado_concluido = False
                st.rerun()

    st.write("")
    renderizar_botoes_navegacao("config")

# -----------------------------------------------------------------------------
# PAINEL DE EXECUÇÃO E RESULTADOS DO SIMULADO
# -----------------------------------------------------------------------------
else:
    if st.session_state.simulado_concluido:
        rolar_para_o_topo()
        total_questoes = len(st.session_state.questoes_online)
        acertos = 0
        for i, q in enumerate(st.session_state.questoes_online):
            if st.session_state.respostas_usuario.get(i) == q["correta"].lower():
                acertos += 1

        erros = total_questoes - acertos
        porcentagem = (acertos / total_questoes * 100) if total_questoes > 0 else 0
        tempo_total = time.time() - st.session_state.tempo_inicio if st.session_state.tempo_inicio else 0
        tempo_str = formatar_tempo(tempo_total)

        st.markdown(
            f"""
            <div class="card-resultado" style="padding: 20px;">
                <h2 style="font-size: 1.6rem; margin-bottom: 8px; color: {text_color};">🏆 SIMULADO CONCLUÍDO!</h2>
                <div style="display: grid; grid-template-columns: repeat(2, 1fr); gap: 10px; margin-top: 15px;">
                    <div class="metric-box">
                        <div style="font-size: 1.4rem; font-weight: bold; color: {text_color};">{acertos} / {total_questoes}</div>
                        <div style="font-size: 0.75rem; text-transform: uppercase; color: {text_color};">Acertos</div>
                    </div>
                    <div class="metric-box">
                        <div style="font-size: 1.4rem; font-weight: bold; color: {text_color};">{erros}</div>
                        <div style="font-size: 0.75rem; text-transform: uppercase; color: {text_color};">Erros</div>
                    </div>
                    <div class="metric-box">
                        <div style="font-size: 1.4rem; font-weight: bold; color: {text_color};">{porcentagem:.1f}%</div>
                        <div style="font-size: 0.75rem; text-transform: uppercase; color: {text_color};">Aproveitamento</div>
                    </div>
                    <div class="metric-box">
                        <div style="font-size: 1.4rem; font-weight: bold; color: {text_color};">⏱️ {tempo_str}</div>
                        <div style="font-size: 0.75rem; text-transform: uppercase; color: {text_color};">Tempo Total</div>
                    </div>
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )

        st.write("")
        
        # Botão de Exportação para Excel integrado na tela de conclusão
        excel_buffer = criar_arquivo_excel(st.session_state.questoes_online)
        st.download_button(
            label="📊 BAIXAR SIMULADO EM EXCEL (.XLSX)",
            data=excel_buffer,
            file_name="Simulado_Rei_Dos_Simulados.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

        st.write("")
        col_res1, col_res2 = st.columns(2)
        with col_res1:
            if st.button("🔄 REINICIAR SIMULADO", key="btn_reiniciar_simulado", use_container_width=True):
                st.session_state.respostas_usuario = {}
                st.session_state.indice_questao = 0
                st.session_state.tempo_inicio = time.time()
                st.session_state.simulado_concluido = False
                st.rerun()

        with col_res2:
            if st.button("⬅️ REVISAR QUESTÕES", key="btn_nav_anterior", use_container_width=True):
                st.session_state.simulado_concluido = False
                st.session_state.indice_questao = total_questoes - 1
                st.rerun()

    else:
        rolar_para_o_topo()
        idx = st.session_state.indice_questao
        questao_atual = st.session_state.questoes_online[idx]
        total_q = len(st.session_state.questoes_online)

        tempo_decorrido = time.time() - st.session_state.tempo_inicio if st.session_state.tempo_inicio else 0
        tempo_fmt = formatar_tempo(tempo_decorrido)

        st.markdown(
            f"""
            <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 8px;">
                <span style="font-weight: 800; font-size: 1rem; color: {text_color};">QUESTÃO {idx + 1} DE {total_q}</span>
                <span style="font-weight: 700; font-size: 0.95rem; color: #EF4444;">⏱️ {tempo_fmt}</span>
            </div>
            """,
            unsafe_allow_html=True,
        )
        st.progress((idx + 1) / total_q)

        enunciado = limpar_e_formatar_texto(questao_atual.get("enunciado", ""))
        st.markdown(f'<div class="enunciado-grande">{enunciado}</div>', unsafe_allow_html=True)

        st.markdown('<div class="rotulo-seletor">Escolha a alternativa correta:</div>', unsafe_allow_html=True)

        resposta_salva = st.session_state.respostas_usuario.get(idx, None)
        alternativas = questao_atual.get("alternativas", {})
        opcao_correta = questao_atual.get("correta", "").lower()

        for chave in sorted(alternativas.keys()):
            letra = chave.lower()
            texto_alt = limpar_e_formatar_texto(alternativas[chave])
            rotulo_botao = f"{letra.upper()}) {texto_alt}"

            if st.button(rotulo_botao, key=f"alt_{idx}_{letra}", use_container_width=True):
                st.session_state.respostas_usuario[idx] = letra
                st.rerun()

        if resposta_salva:
            st.write("")
            if resposta_salva == opcao_correta:
                st.success(f"✅ Resposta Correta! A alternativa certa é a ({opcao_correta.upper()}).")
            else:
                st.error(f"❌ Resposta Incorreta. A alternativa correta era a ({opcao_correta.upper()}).")

            explicacao = limpar_e_formatar_texto(questao_atual.get("explicacao", ""))
            st.info(f"💡 **Explicação:** {explicacao}")

        st.write("")
        col_ant, col_prox = st.columns(2)
        with col_ant:
            if idx > 0:
                if st.button("⬅️ Questão Anterior", use_container_width=True):
                    st.session_state.indice_questao -= 1
                    st.rerun()
        with col_prox:
            if idx < total_q - 1:
                if st.button("Próxima Questão ➡️", use_container_width=True):
                    st.session_state.indice_questao += 1
                    st.rerun()
            else:
                if st.button("🏁 Finalizar Simulado", use_container_width=True):
                    st.session_state.simulado_concluido = True
                    st.rerun()
